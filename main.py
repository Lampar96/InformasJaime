import datetime
import pandas as pd
import os
import argparse
from openpyxl import load_workbook

parser = argparse.ArgumentParser(description="Cuenta tickets serviceNow")
parser.add_argument("-p", "--path", dest="path", type=str, help="Ruta de descarga", required=True, metavar="")


def rellenarExcel(dest):
    wb = load_workbook('Libro1.xlsx')
    ws = wb['Hoja1']
    df = pd.read_excel('Libro1.xlsx', 'Hoja1')
    num_filas = df.shape[0]
    for row in range(1, num_filas):
        if str(ws.cell(row, 1).value)[0:7] == str(datetime.datetime.now())[0:7]:
            for path in os.listdir(dest):
                if "InformesLBID" in path:
                    df = pd.read_excel(path)
                    row_num = df.shape[0]
                    if path[13:14] == '+':
                        cell = path[14:15] + str(row)
                        if ws[cell].value == None:
                            num = row_num
                        else:
                            num = ws[cell].value + row_num
                        ws[cell] = num
                    else:
                        cell = path[13:14] + str(row)
                        ws[cell] = row_num
                    os.remove(dest + '/' + path)
            wb.save('Libro1.xlsx')
            break

if __name__ == "__main__":
    args = parser.parse_args()
    rellenarExcel(args.path)
